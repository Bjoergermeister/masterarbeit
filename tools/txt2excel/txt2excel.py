import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

from data_importer import import_data
from data_importer import benchmarks

COLUMNS = []
MODES = ["Regular", "Manual", "Container"]
for i in range(0, 52):
    prefix = 'A' if i > 25 else ''
    ascii_code = i if i <= 25 else i - 26
    COLUMNS.append("{}{}".format(prefix, chr(ascii_code + 65)))


OPERATION_CASES_COUNTS = {
    "Open": 16,
    "Read": 16,
    "Write": 6,
    "Create": 1,
    "Delete": 6
}

CASE_LABELS = {
    "Open": range(0, 16),
    "Read": range(0, 16),
    "Write": ["1", "128", "256", "512", "1024", "2048"],
    "Delete": ["1", "128", "256", "512", "1024", "2048"]
}

def get_columns_for_mode(operation, mode):
    mode_index = MODES.index(mode)
    multiplier = 2 if operation == "Write" else 1
    start_index = (mode_index * OPERATION_CASES_COUNTS[operation] * multiplier + (1 * mode_index))
    end_index = start_index + OPERATION_CASES_COUNTS[operation] * multiplier
    return COLUMNS[start_index:end_index]

def get_case_label_for_index(operation, index):
    if operation == "Write":
        case = str(CASE_LABELS[operation][math.ceil(index // 2)])
        return f"{case} (2)" if index % 2 == 1 else case
    else:
        return str(CASE_LABELS[operation][index])
    
def get_values(operation, language, mode, case_index):
    if operation == "Create":
        return benchmarks["Filesystem"][operation][language][mode]
    else:
        case = get_case_label_for_index(operation, case_index)
        if case.endswith(")"):
            case = case[:-4]
        return benchmarks["Filesystem"][operation][language][mode][case]


def add_spreadsheet_headers(workbook):
    sheets = ["Open", "Read", "Write", "Create", "Delete"]
    for sheet in sheets:
        add_spreadsheet_header(workbook[sheet], sheet)

def add_spreadsheet_header(sheet, operation):
    for i in range(0, 3):
        multiplier = 2 if operation == "Write" else 1
        range_start = (i * OPERATION_CASES_COUNTS[operation] * multiplier + (1 * i))
        range_end = range_start + OPERATION_CASES_COUNTS[operation] * multiplier - 1
        
        sheet.merge_cells(start_row=1, start_column=range_start + 1, end_row=1, end_column=range_end + 1)
        column_code = f"{COLUMNS[range_start]}1" 
        sheet[column_code] = MODES[i]
        sheet[column_code].alignment = Alignment(horizontal="center")
        sheet[column_code].fill = PatternFill(fgColor="cccccc", fill_type="solid")
        
        if operation == "Create":
            continue

        for index, column in enumerate(range(range_start, range_end + 1)):
            column_code = f"{COLUMNS[column]}2"
            label = get_case_label_for_index(operation, index) 
            sheet[column_code] = str(int(label) + 1) if operation in ["Read", "Open"] else label 
            sheet[column_code].alignment = Alignment(horizontal="center")
            sheet[column_code].fill = PatternFill(fgColor="cccccc", fill_type="solid")


def add_data_to_spreadsheet(workbook, language):
    for operation in list(benchmarks["Filesystem"].keys()):
        sheet = workbook[operation]

        # Regular
        start_row = 2 if operation == "Create" else 3
        for mode in MODES:
            for index, column in enumerate(get_columns_for_mode(operation, mode)):
                values = get_values(operation, language, mode, index)
                for row in range(0, 100):
                    cell = f"{column}{start_row + row}"
                    sheet[cell] = values[row][index % 2] if operation == "Write" else values[row]
        
        # Add formulas
        MATH_OPERATIONS = ["MIN", "MAX", "SUM"]
        DIVIDERS = ["", "", "/100"]
        ALL_COLUMNS = [get_columns_for_mode(operation, mode) for mode in MODES]
        for index, mode in enumerate(MODES):
            for list_index, column_list in enumerate(ALL_COLUMNS):
                for column_index, value_column in enumerate(column_list):
                    formula = f"={MATH_OPERATIONS[index]}({value_column}{start_row}:{value_column}{start_row + 99}){DIVIDERS[index]}"
                    start_row_offset = 101 + list_index
                    sheet[f"{ALL_COLUMNS[index][column_index]}{start_row + start_row_offset}"] = formula

def prepare_workbook(language, filename):
    workbook = Workbook()

    # Rename sheets
    workbook.active.title = "Open"
    workbook.create_sheet("Read", 2)
    workbook.create_sheet("Write", 3)
    workbook.create_sheet("Create", 4)
    workbook.create_sheet("Delete", 5)

    add_spreadsheet_headers(workbook)
    add_data_to_spreadsheet(workbook, language)
    workbook.save(filename=filename)

if __name__ == "__main__":
    import_data("../../Results", benchmarks)
    prepare_workbook("C", "../../Filesystem_C.xlsx")
    prepare_workbook("Java", "../../Filesystem_Java.xlsx")